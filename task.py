
import re
import os
import xlwt
import urllib
import openpyxl
import webbrowser
import pandas as pd
from numpy import datetime_data
import urllib.request, urllib.error

# 河南的市：郑州、开封、洛阳、平顶山、安阳、鹤壁、新乡、焦作、濮阳、许昌、漯河、三门峡、商丘、周口、驻马店、南阳、信阳
HENAN = ['ZHENGZHOU', 'KAIFENG', 'LUOYANG', 'PINGDINGSHAN', 'ANYANG', 'HEBI', 'XINXIANG', 'JIAOZUO', 'PUYANG', 'XUCHANG', 'LUOHE', 'SANMENXIA', 'SHANGQIU', 'ZHOUKOU', 'ZHUMADIAN', 'NANYANG', 'XINYANG']
# 安徽的市：合肥、巢湖、芜湖、蚌埠、淮南、马鞍山、淮北、铜陵、安庆、黄山、滁州、阜阳、宿州、六安、亳州、池州、宣城
ANHUI = ['HEFEI', 'CHAOHU', 'WUHU', 'BENGBU', 'HUAINAN', 'MAANSHAN', 'HUAIBEI', 'TONGLING', 'ANQING', 'HUANGSHAN', 'CHUZHOU', 'FUYANG', 'SUZHOU', 'LIUAN', 'HAOZHOU', 'CHIZHOU', 'XUANCHENG']
if __name__ == '__main__':
    data = []
    with open('ghcnd-stations.txt', "r") as f:
        data = f.readlines()
    # print(type(city2id))
    # print(type(city2id[0]))
    city2Id = {}
    for myInfo in data:
        lis = myInfo.split()
        if 'CHM' == lis[0][:3]: # 中国
            if lis[4] not in city2Id:
                city2Id[lis[4]] = []
            city2Id[lis[4]].append(lis[-1])
            # print(' '.join(lis))
    # print(city2Id)

    dataAll = []

    for city in HENAN:
        if city in city2Id:
            dataAll.append([])
            dataAll[-1].append('HENAN')
            dataAll[-1].append(city)
            dataAll[-1].append('CHM000' + city2Id[city][0])
            # print(city, city2Id[city])
    for city in ANHUI:
        if city in city2Id:
            dataAll.append([])
            dataAll[-1].append('ANHUI')
            dataAll[-1].append(city)
            dataAll[-1].append('CHM000' + city2Id[city][0])
            # print(city, city2Id[city])
    
    '''
    # 从noaa下载数据
    baseUrl1 = 'https://www.ncei.noaa.gov/pub/data/ghcn/daily/by_station/'
    for i in range(len(dataAll)):
        cityId = dataAll[i][2]
        baseUrl2 = cityId + '.csv.gz'
        url = baseUrl1 + baseUrl2
        print(url)
        urllib.request.urlretrieve(url)
        webbrowser.open(url)
    '''

    '''
    # 解压并保存到项目所在目录,删除解压文件夹(根据解压软件和下载路径不同需要对命令进行修改)
    for i in range(len(dataAll)):
        cityId = dataAll[i][2]
        oss ='D:\\7-Zip\\7z x C:\\Users\\xuzikang\\Downloads\\' + cityId + '.csv.gz' +  r' -o' + cityId
        os.system(oss)

        oss = 'move .\\' + cityId + '\\' + cityId + '.csv .\\'
        os.system(oss)

        oss = 'rd/s/q ' + cityId
        os.system(oss)
    '''
    # 读取和整理爬下来的数据
    
    dataId = []
    dataCity = []
    dataProvince = []
    dataDate = []
    dataTMAX = []
    dataTMIN = []
    dataPRCP = []
    dataTAVG = []
    
    pos = dict()
    pos['TMAX'] = 0
    pos['TMIN'] = 1
    pos['PRCP'] = 2
    pos['TAVG'] = 3
    for i in range(len(dataAll)):
        cityId = dataAll[i][2]
        path = cityId + '.csv'
        df = pd.DataFrame(pd.read_csv(path))
        mydic = {}
        
        for j in range(len(df.index)):
            date = df.iloc[j, 1]
            valType = df.iloc[j, 2]
            val = df.iloc[j, 3]
            if date not in mydic:
                mydic[date] = ['nan', 'nan', 'nan', 'nan']
            if valType not in pos:
                continue
            mydic[date][pos[valType]] = val
        
        for key, value in mydic.items():
            dataId.append(cityId)
            dataCity.append(dataAll[i][1])
            dataProvince.append(dataAll[i][2])
            dataDate.append(key)
            dataTMAX.append(value[0])
            dataTMIN.append(value[1])
            dataPRCP.append(value[2])
            dataTAVG.append(value[3])
    

    # 输出到表格里,用xlwt会超出最大行数，只能使用openpyxl
    book = openpyxl.Workbook()
    sheet = book.create_sheet('climateData', index=0)
    
    col = ('编号', '城市', '省份', '日期', 'TMAX', 'TMIN', 'PRCP', 'TAVG')
    for i in range(8):
        sheet.cell(1, i + 1, col[i])

    for i in range(len(dataId)):
        sheet.cell(i + 2, 1, dataId[i])
        sheet.cell(i + 2, 2, dataCity[i])
        sheet.cell(i + 2, 3, dataProvince[i])
        # print(dataDate[i])
        # print(type(dataDate[i]))
        sheet.cell(i + 2, 4, str(dataDate[i]))
        sheet.cell(i + 2, 5, str(dataTMAX[i])) 
        sheet.cell(i + 2, 6, str(dataTMIN[i]))
        sheet.cell(i + 2, 7, str(dataPRCP[i]))
        sheet.cell(i + 2, 8, str(dataTAVG[i]))

    book.save('climateData.csv')